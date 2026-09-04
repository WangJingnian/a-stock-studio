# -*- coding: utf-8 -*-
"""同花顺投资账本同步服务：把账本「汇总持仓 + 交易流水」导入本地账户。

流程：
1. 拉取账本全部账户（account_list）
2. 逐个账户拉持仓（stock_position）与交易流水（get_money_history）
3. 汇总：持仓按代码加权合并，交易合并去重
4. 导入本地账户（自动创建「同花顺汇总」账户，幂等可重复执行）
5. 用账本当前持仓做校正，保证本地持仓与账本一致
6. 可选：导入资产历史曲线（asset_trend）到本地日快照
"""
import hashlib
import time
import uuid
import json
import os
from datetime import date, timedelta
from typing import Any, Dict, List, Optional

from src.repositories.portfolio_repo import PortfolioRepository
from src.services.portfolio_service import PortfolioService
from src.services.ths_sync.ths_client import ThsSession

DEFAULT_ACCOUNT_NAME = "百福具臻"

# 窗口覆盖：导入交易流水时只重写最近 N 个交易日，更早历史保留
WINDOW_TRADING_DAYS = 7


def _n_trading_days_before(d: date, n: int) -> date:
    """返回 ``d`` 往前第 ``n`` 个交易日（跳过周末；法定节假日未单独建模，
    窗口取 7 个交易日已含一定余量）。"""
    cur = d
    counted = 0
    while counted < n:
        cur -= timedelta(days=1)
        if cur.weekday() < 5:  # 周一至周五
            counted += 1
    return cur


def _dedup_key(t: Dict[str, Any]) -> str:
    raw = "|".join([
        str(t.get("code", "")),
        str(t.get("trade_date", "")),
        str(t.get("side", "")),
        str(t.get("quantity", "")),
        str(t.get("price", "")),
        str(t.get("trans_no", "")),
    ])
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def _skip_symbol(code: str, name: str = "") -> bool:
    """过滤非持仓标的：国债逆回购等。"""
    c = (code or "").strip()
    n = (name or "").strip()
    if not c:
        return True
    # 国债逆回购 204xxx / 1318xx / 1319xx
    if c.startswith("204") or c.startswith("1318") or c.startswith("1319"):
        return True
    if "逆回购" in n or n.startswith("GC") or n.startswith("R-"):
        return True
    return False


def _prev_mv_of(p: Dict[str, Any]) -> float:
    """根据当日盈亏额/率反推昨收市值：昨收 = 今市值 / (1 + 当日盈亏率)。"""
    value = float(p.get("value") or 0.0)
    rate = float(p.get("day_pnl_pct") or 0.0)
    if abs(rate) > 1e-12:
        return value / (1.0 + rate)
    return value


# 场外基金过滤：仅保留股票 + 场内 ETF。
# - 名称含基金类型词（混合/债券/LOF/联接/货币/纯债/指数C 等）
# - 代码落在场外基金区间（01-09 / 16 开头；16 开头为 LOF，160514 债券类）
# 注意 000/003 开头既有深市股票也有场外基金（如 000480 东方红新动力混合A），
# 此类依赖名称关键词识别，防止误杀 000630/000999/003816 等股票。
_FUND_NAME_KEYWORDS = (
    "混合",
    "债券",
    "LOF",
    "联接",
    "货币",
    "纯债",
    "信用债",
    "指数C",
    "ETF联接",
)
_FUND_CODE_PREFIXES = ("16", "01", "02", "03", "04", "05", "06", "07", "08", "09")


def _is_fund_symbol(code: str, name: str = "") -> bool:
    """判断是否为场外基金（非股票/场内 ETF）。"""
    c = (code or "").strip()
    n = (name or "").strip()
    if any(k in n for k in _FUND_NAME_KEYWORDS):
        return True
    if c.startswith(_FUND_CODE_PREFIXES):
        return True
    return False


class ThsSyncService:
    def __init__(
        self,
        *,
        cookie_file: Optional[str] = None,
        portfolio_service: Optional[PortfolioService] = None,
        repo: Optional[PortfolioRepository] = None,
    ):
        data_dir = os.path.join(
            os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))),
            "data",
        )
        os.makedirs(data_dir, exist_ok=True)
        self.cookie_file = cookie_file or os.path.join(data_dir, "ths_cookies.json")
        self.client = ThsSession(self.cookie_file)
        self.portfolio_service = portfolio_service or PortfolioService()
        self.repo = repo or PortfolioRepository()

    # ------------------------------------------------------------------
    # 登录
    # ------------------------------------------------------------------
    def get_status(self) -> Dict[str, Any]:
        logged = self.client.is_logged_in()
        return {
            "logged_in": logged,
            "platform": "同花顺投资账本",
            "cookie_file": os.path.basename(self.cookie_file),
        }

    def create_qrcode(self) -> Dict[str, Any]:
        return self.client.create_qrcode()

    def poll_login(self, qrid: str, timeout: float = 180.0) -> Dict[str, Any]:
        ok = self.client.poll_login(qrid, timeout=timeout)
        return {"success": ok, "logged_in": ok}

    def logout(self) -> Dict[str, Any]:
        self.client.clear_login()
        return {"logged_in": False}

    # ------------------------------------------------------------------
    # 汇总数据拉取
    # ------------------------------------------------------------------
    def fetch_merged(self, start_date: Optional[str] = None, end_date: Optional[str] = None) -> Dict[str, Any]:
        """拉取账本全部账户并汇总持仓/交易/现金。start_date/end_date 为 YYYYMMDD 交易时间筛选。"""
        accounts = self.client.fetch_accounts()
        merged_positions: Dict[str, Dict[str, Any]] = {}
        all_trades: List[Dict[str, Any]] = []
        total_cash = 0.0
        account_meta: List[Dict[str, Any]] = []

        for acc in accounts:
            if acc["type"] != "common":
                continue
            fk = acc["fund_key"]
            account_meta.append({"fund_key": fk, "name": acc["name"]})
            try:
                pos = self.client.fetch_positions(fk)
            except Exception:  # noqa: BLE001
                continue
            total_cash += pos.get("money_remain", 0.0)
            for p in pos.get("positions", []):
                code = str(p.get("code"))
                if not code or _skip_symbol(code, str(p.get("name") or "")):
                    continue
                if code in merged_positions:
                    m = merged_positions[code]
                    q1 = m["quantity"]
                    q2 = p.get("quantity", 0.0)
                    if q1 + q2 > 0:
                        m["cost"] = (m["cost"] * q1 + p.get("cost", 0.0) * q2) / (q1 + q2)
                    m["quantity"] = q1 + q2
                    m["value"] = m.get("value", 0.0) + p.get("value", 0.0)
                    m["hold_profit"] = m.get("hold_profit", 0.0) + p.get("hold_profit", 0.0)
                    m["day_pnl"] = m.get("day_pnl", 0.0) + p.get("day_pnl", 0.0)
                    m["prev_mv"] = m.get("prev_mv", 0.0) + _prev_mv_of(p)
                    m["hold_days"] = max(int(m.get("hold_days") or 0), int(p.get("hold_days") or 0))
                    m["account_count"] += 1
                else:
                    merged_positions[code] = {
                        **p,
                        "code": code,
                        "account_count": 1,
                        "prev_mv": _prev_mv_of(p),
                    }
            # 交易流水（分页拉取，最多 20 页）
            page = 1
            while page <= 20:
                try:
                    trades = self.client.fetch_trades(
                        fk, page=page, count=100, start_date=start_date, end_date=end_date
                    )
                except Exception:  # noqa: BLE001
                    break
                if not trades:
                    break
                all_trades.extend(trades)
                if len(trades) < 100:
                    break
                page += 1

        # 计算合并后的当日盈亏率（当日盈亏 / 昨收市值）
        for m in merged_positions.values():
            prev_mv = float(m.get("prev_mv") or 0.0)
            day_pnl = float(m.get("day_pnl") or 0.0)
            if prev_mv > 0:
                m["day_pnl_pct"] = day_pnl / prev_mv
            else:
                m["day_pnl_pct"] = 0.0

        return {
            "accounts": account_meta,
            "positions": list(merged_positions.values()),
            "trades": all_trades,
            "total_cash": round(total_cash, 2),
            "position_count": len(merged_positions),
            "trade_count": len(all_trades),
        }

    def list_merged_trades(self, start_date: Optional[str] = None, end_date: Optional[str] = None) -> Dict[str, Any]:
        """从账本拉取全部账户交易流水（可选时间范围），合并去重后返回（不写入本地成本）。"""
        accounts = self.client.fetch_accounts()
        all_trades: List[Dict[str, Any]] = []
        for acc in accounts:
            if acc["type"] != "common":
                continue
            page = 1
            while page <= 20:
                try:
                    trades = self.client.fetch_trades(
                        acc["fund_key"], page=page, count=100, start_date=start_date, end_date=end_date
                    )
                except Exception:  # noqa: BLE001
                    break
                if not trades:
                    break
                all_trades.extend(trades)
                if len(trades) < 100:
                    break
                page += 1
        # 按 trans_no 去重
        seen = set()
        unique = []
        for t in all_trades:
            key = str(t.get("trans_no") or "") or _dedup_key(t)
            if key in seen:
                continue
            seen.add(key)
            unique.append(t)
        unique.sort(key=lambda x: (str(x.get("trade_date") or ""), str(x.get("trade_time") or "")), reverse=True)
        return {"total": len(unique), "trades": unique}

    def list_local_import_records(
        self,
        *,
        start_date: Optional[str] = None,
        end_date: Optional[str] = None,
        account_name: str = DEFAULT_ACCOUNT_NAME,
    ) -> Dict[str, Any]:
        """查询本地导入的账本导出流水（含逆回购/分红/银证转账等全部类别）。

        账本实时接口（stock_transaction）不返回逆回购等部分记录，
        导出文件「交易记录」sheet 全类别持久化到 portfolio_import_records，
        本方法读取该表，按日期范围过滤、最新在前返回。
        """
        account = self._find_or_create_account(account_name)
        d0 = date.fromisoformat(start_date) if start_date else date(2000, 1, 1)
        d1 = date.fromisoformat(end_date) if end_date else date.today()
        if d1 < d0:
            d0, d1 = d1, d0
        rows = self.repo.list_import_records(account_id=account["id"], start_date=d0, end_date=d1)
        rows.sort(key=lambda r: (str(r.trade_date), str(r.trade_time or "")), reverse=True)
        records = []
        for r in rows:
            records.append({
                "code": r.symbol or "",
                "name": r.name or "",
                "record_type": r.record_type,
                "trade_date": r.trade_date.isoformat() if hasattr(r.trade_date, "isoformat") else str(r.trade_date),
                "trade_time": r.trade_time or "",
                "quantity": float(r.quantity or 0),
                "price": float(r.price or 0),
                "amount": float(r.amount or 0),
                "fee": float(r.fee or 0),
                "note": r.note or "",
            })
        return {"total": len(records), "records": records}


    def stock_ledger(self, *, start_date: str = "", end_date: str = "") -> Dict[str, Any]:
        """按股票维度汇总本地导入流水（买入/卖出/分红/税费），返回股票列表与各自明细。

        数据源为账本导出文件导入的 portfolio_import_records（含分红/股息税等全类别）。
        银证转账等资金往来不计入个股流水。返回 stocks 按「最近有记录的在前」排序。
        """
        d0 = start_date or ""
        d1 = end_date or ""
        res = self.list_local_import_records(start_date=d0 or None, end_date=d1 or None)
        groups: Dict[str, Dict[str, Any]] = {}

        def _g(symbol: str, name: str) -> Dict[str, Any]:
            if symbol not in groups:
                groups[symbol] = {
                    "symbol": symbol,
                    "name": name or "",
                    "buy_count": 0, "buy_amount": 0.0, "buy_fee": 0.0,
                    "sell_count": 0, "sell_amount": 0.0, "sell_fee": 0.0,
                    "dividend_count": 0, "dividend_amount": 0.0,
                    "adjust_count": 0, "adjust_amount": 0.0,
                    "other_fee": 0.0,
                    "records": [],
                    "latest_date": "",
                }
            return groups[symbol]

        for r in res.get("records", []):
            cat = str(r.get("record_type") or "").strip()
            symbol = str(r.get("code") or "").strip()
            if not symbol:
                continue
            name = str(r.get("name") or "")
            if _skip_symbol(symbol, name):
                continue  # 国债逆回购等非持仓标的
            if cat in self._EXPORT_CATEGORY_CASH_IN or cat in self._EXPORT_CATEGORY_CASH_OUT:
                continue  # 银证转账等资金往来不计入个股
            g = _g(symbol, name)
            amt = float(r.get("amount") or 0)
            fee = float(r.get("fee") or 0)
            d = str(r.get("trade_date") or "")
            if d and d > g["latest_date"]:
                g["latest_date"] = d
            g["records"].append({
                "record_type": cat,
                "date": d,
                "time": str(r.get("trade_time") or ""),
                "quantity": float(r.get("quantity") or 0),
                "price": float(r.get("price") or 0),
                "amount": amt,
                "fee": fee,
                "note": r.get("note") or "",
            })
            if cat == "买入":
                g["buy_count"] += 1
                g["buy_amount"] += abs(amt)
                g["buy_fee"] += fee
            elif cat == "卖出":
                g["sell_count"] += 1
                g["sell_amount"] += abs(amt)
                g["sell_fee"] += fee
            elif cat == "除权除息":
                if amt > 0:
                    g["dividend_count"] += 1
                    g["dividend_amount"] += amt
                else:
                    g["adjust_count"] += 1
                    g["adjust_amount"] += abs(amt)
            elif "股息个税" in cat or cat == "缴税":
                g["other_fee"] += abs(amt)
            # 其他类别（利息等）仅入明细，不参与上述聚合

        stocks = list(groups.values())
        stocks.sort(key=lambda x: x["latest_date"], reverse=True)
        for st in stocks:
            st["records"].sort(key=lambda x: (x["date"], x["time"]), reverse=True)
            st["buy_amount"] = round(st["buy_amount"], 2)
            st["sell_amount"] = round(st["sell_amount"], 2)
            st["buy_fee"] = round(st["buy_fee"], 2)
            st["sell_fee"] = round(st["sell_fee"], 2)
            st["dividend_amount"] = round(st["dividend_amount"], 2)
            st["adjust_amount"] = round(st["adjust_amount"], 2)
            st["other_fee"] = round(st["other_fee"], 2)
        return {"total_stocks": len(stocks), "stocks": stocks}


    def holding_ledger(self, *, start_date: str = "", end_date: str = "") -> Dict[str, Any]:
        """当前持仓表格数据：账本持仓（含现成 hold_days/成本/现价/盈亏）+ 本地导入流水明细。

        持仓字段优先取账本实时接口（hold_days 为账本现成值，无需自行计算）；
        未登录时回退本地快照（此时持仓天数为 0，前端显示 --）。
        """
        if self.client.is_logged_in():
            data = self.fetch_merged()
            positions = data.get("positions", [])
        else:
            positions = []
            account = self._find_or_create_account(DEFAULT_ACCOUNT_NAME)
            snapshot = self.portfolio_service.get_portfolio_snapshot(account_id=account["id"], include_realtime=False)
            for acc in snapshot.get("accounts", []):
                for p in acc.get("positions", []):
                    positions.append({
                        "code": str(p.get("symbol") or ""),
                        "name": str(p.get("name") or ""),
                        "quantity": float(p.get("quantity") or 0),
                        "cost": float(p.get("avg_cost") or 0),
                        "price": float(p.get("last_price") or 0),
                        "hold_profit": float(p.get("unrealized_pnl_base") or 0),
                        "hold_rate": float(p.get("unrealized_pnl_pct") or 0),
                        "hold_days": 0,
                        "day_pnl": float(p.get("day_pnl") or 0),
                        "day_pnl_pct": float(p.get("day_pnl_pct") or 0),
                    })

        # 本地导入流水按 code 索引（含分红/股息税等全类别，过滤逆回购与银证转账）
        res = self.list_local_import_records(start_date=start_date or None, end_date=end_date or None)
        by_code: Dict[str, List[Dict[str, Any]]] = {}
        for r in res.get("records", []):
            code = str(r.get("code") or "").strip()
            if not code:
                continue
            name = str(r.get("name") or "")
            if _skip_symbol(code, name):
                continue
            cat = str(r.get("record_type") or "").strip()
            if cat in self._EXPORT_CATEGORY_CASH_IN or cat in self._EXPORT_CATEGORY_CASH_OUT:
                continue
            by_code.setdefault(code, []).append(r)

        stocks: List[Dict[str, Any]] = []
        for p in positions:
            code = str(p.get("code") or "")
            qty = float(p.get("quantity") or 0)
            if not code or qty <= 0:
                continue
            name = str(p.get("name") or "")
            buy_c = buy_a = buy_f = sell_c = sell_a = sell_f = 0.0
            div_c = div_a = adj_c = adj_a = other_f = 0.0
            buy_c = sell_c = div_c = adj_c = 0
            recs: List[Dict[str, Any]] = []
            for r in sorted(by_code.get(code, []), key=lambda x: (str(x.get("trade_date") or ""), str(x.get("trade_time") or "")), reverse=True):
                cat = str(r.get("record_type") or "").strip()
                amt = float(r.get("amount") or 0)
                fee = float(r.get("fee") or 0)
                recs.append({
                    "record_type": cat,
                    "date": str(r.get("trade_date") or ""),
                    "time": str(r.get("trade_time") or ""),
                    "quantity": float(r.get("quantity") or 0),
                    "price": float(r.get("price") or 0),
                    "amount": amt,
                    "fee": fee,
                    "note": r.get("note") or "",
                })
                if cat == "买入":
                    buy_c += 1; buy_a += abs(amt); buy_f += fee
                elif cat == "卖出":
                    sell_c += 1; sell_a += abs(amt); sell_f += fee
                elif cat == "除权除息":
                    if amt > 0:
                        div_c += 1; div_a += amt
                    else:
                        adj_c += 1; adj_a += abs(amt)
                elif "股息个税" in cat or cat == "缴税":
                    other_f += abs(amt)
            stocks.append({
                "symbol": code,
                "name": name,
                "quantity": qty,
                "cost": float(p.get("cost") or 0),
                "last_price": float(p.get("price") or p.get("last_price") or 0),
                "hold_profit": float(p.get("hold_profit") or 0),
                "hold_rate": float(p.get("hold_rate") or 0),
                "hold_days": int(p.get("hold_days") or 0),
                "day_pnl": float(p.get("day_pnl") or 0),
                "day_pnl_pct": float(p.get("day_pnl_pct") or 0),
                "buy_count": buy_c, "buy_amount": round(buy_a, 2), "buy_fee": round(buy_f, 2),
                "sell_count": sell_c, "sell_amount": round(sell_a, 2), "sell_fee": round(sell_f, 2),
                "dividend_count": div_c, "dividend_amount": round(div_a, 2),
                "adjust_count": adj_c, "adjust_amount": round(adj_a, 2),
                "other_fee": round(other_f, 2),
                "records": recs,
            })
        stocks.sort(key=lambda x: (x["hold_days"], x["symbol"]), reverse=True)
        return {"total": len(stocks), "stocks": stocks}


    def build_statement(self, *, month: str, account_id: Optional[int] = None, cost_method: str = "fifo") -> Dict[str, Any]:
        """从账本实时拉取指定月份的全部交易（含国债逆回购等），聚合生成月度对账单。

        资产期初/期末仍使用本地每日快照计算。不写入本地持仓成本。
        """
        import calendar

        try:
            y, m = (int(p) for p in str(month).split("-"))
        except Exception:
            raise ValueError("month must be YYYY-MM")
        if m < 1 or m > 12:
            raise ValueError("month must be YYYY-MM")
        last_day = calendar.monthrange(y, m)[1]
        start_date = f"{y}{m:02d}01"
        end_date = f"{y}{m:02d}{last_day:02d}"
        date_from = date(y, m, 1)
        next_month = date(y + 1, 1, 1) if m == 12 else date(y, m + 1, 1)
        date_to = next_month - timedelta(days=1)
        if account_id is None:
            # 缺省账户：优先同花顺汇总（含导入流水），保证基于导出数据的聚合可用
            account_id = self._find_or_create_account(DEFAULT_ACCOUNT_NAME)["id"]

        buy_count = 0
        buy_amount = 0.0
        buy_fee = 0.0
        sell_count = 0
        sell_amount = 0.0
        sell_fee = 0.0
        cash_in = 0.0
        cash_out = 0.0
        details: List[Dict[str, Any]] = []
        dividends: List[Dict[str, Any]] = []

        # 优先使用本地导入的导出流水（含分红/银证转账等完整类别）
        imported = []
        if account_id:
            imported = self.repo.list_import_records(account_id, date_from, date_to)

        if imported:
            for t in imported:
                cat = (t.record_type or "").strip()
                d = t.trade_date.isoformat() if t.trade_date else ""
                tm = str(t.trade_time or "")
                qty = float(t.quantity or 0)
                px = float(t.price or 0)
                amt = float(t.amount or 0)
                fee = float(t.fee or 0)
                symbol = str(t.symbol or "")
                name = str(t.name or "")
                if cat in ("买入", "卖出"):
                    # 导出「发生金额」：买入为负（资金流出），卖出为正（资金流入）
                    amt_abs = abs(amt)
                    details.append({
                        "date": d,
                        "time": tm,
                        "symbol": symbol,
                        "name": name,
                        "side": "buy" if cat == "买入" else "sell",
                        "quantity": qty,
                        "price": px,
                        "amount": round(amt_abs, 2),
                        "fee": round(fee, 2),
                    })
                    if cat == "买入":
                        buy_count += 1
                        buy_amount += amt_abs
                        buy_fee += fee
                    else:
                        sell_count += 1
                        sell_amount += amt_abs
                        sell_fee += fee
                elif cat == "除权除息":
                    # 真正的分红/除权记录：正额为分红入账，负额为除权调整
                    if amt > 0:
                        dividends.append({"date": d, "symbol": symbol, "name": name, "amount": amt, "type": "分红"})
                    else:
                        dividends.append({"date": d, "symbol": symbol, "name": name, "amount": amt, "type": "除权调整"})
                elif "股息个税" in cat or cat == "缴税":
                    # 股息个税：持股不满一年卖出补扣的红利税，属卖出税费，不计入分红
                    sell_fee += abs(amt)
                elif cat in self._EXPORT_CATEGORY_CASH_IN:
                    cash_in += abs(amt)
                elif cat in self._EXPORT_CATEGORY_CASH_OUT:
                    cash_out += abs(amt)
        else:
            # 回退：实时账本拉取（无导入流水时）
            res = self.list_merged_trades(start_date=start_date, end_date=end_date)
            trades = res.get("trades", [])
            for t in trades:
                qty = float(t.get("quantity") or 0)
                px = float(t.get("price") or 0)
                amt = qty * px
                fee = float(t.get("fee") or 0)
                side = (t.get("side") or "buy").strip().lower()
                op_name = str(t.get("op_name") or "")
                # 过滤数量为 0 的无效记录（账本同步可能带入空交易噪音）
                if qty <= 0:
                    continue
                details.append({
                    "date": str(t.get("trade_date") or ""),
                    "time": str(t.get("trade_time") or ""),
                    "symbol": str(t.get("code") or ""),
                    "name": str(t.get("name") or ""),
                    "side": "buy" if side == "buy" else "sell",
                    "quantity": qty,
                    "price": px,
                    "amount": round(amt, 2),
                    "fee": round(fee, 2),
                })
                if side == "buy":
                    buy_count += 1
                    buy_amount += amt
                    buy_fee += fee
                else:
                    sell_count += 1
                    sell_amount += amt
                    sell_fee += fee

        snapshots = self.repo.list_daily_snapshots_for_risk(
            as_of=date_to,
            cost_method="fifo",
            account_id=account_id,
            lookback_days=62,
        )
        end_snap = snapshots[-1] if snapshots else None
        begin_snap: Any = None
        for s in snapshots:
            if s.snapshot_date < date_from:
                begin_snap = s

        def _eq(snap: Any) -> Optional[float]:
            return float(snap.total_equity or 0) if snap is not None else None

        begin_equity = _eq(begin_snap)
        end_equity = _eq(end_snap)
        ret_pct = None
        if begin_equity and begin_equity != 0 and end_equity is not None:
            ret_pct = round((end_equity - begin_equity) / begin_equity * 100, 2)

        return {
            "month": month,
            "source": "ths",
            "trades": {
                "buy_count": buy_count,
                "buy_amount": round(buy_amount, 2),
                "buy_fee": round(buy_fee, 2),
                "sell_count": sell_count,
                "sell_amount": round(sell_amount, 2),
                "sell_fee": round(sell_fee, 2),
                "net_cash_outflow": round(buy_amount + buy_fee - sell_amount, 2),
            },
            "cash": {"inflow": round(cash_in, 2), "outflow": round(cash_out, 2), "net": round(cash_in - cash_out, 2)},
            "dividends": {
                "count": len(dividends),
                "items": sorted(
                    dividends,
                    key=lambda d: (str(d.get("date") or ""), str(d.get("symbol") or "")),
                    reverse=True,
                ),
            },
            "asset": {"begin_equity": begin_equity, "end_equity": end_equity, "return_pct": ret_pct},
            "details": sorted(
                details,
                key=lambda d: (str(d.get("date") or ""), str(d.get("time") or "")),
                reverse=True,
            ),
        }

    def fetch_asset_trend_all(self) -> List[Dict[str, Any]]:
        """汇总各账户的资产历史曲线（按日期合并总资产）。"""
        accounts = self.client.fetch_accounts()
        merged: Dict[str, Dict[str, Any]] = {}
        for acc in accounts:
            if acc["type"] != "common":
                continue
            try:
                points = self.client.fetch_asset_trend(acc["fund_key"])
            except Exception:  # noqa: BLE001
                continue
            for p in points:
                d = str(p.get("date") or "")
                if not d:
                    continue
                if d not in merged:
                    merged[d] = {"date": d, "asset": 0.0, "profit": 0.0}
                merged[d]["asset"] += p.get("asset", 0.0)
                merged[d]["profit"] += p.get("profit", 0.0)
        out = list(merged.values())
        out.sort(key=lambda x: x["date"])
        return out

    # ------------------------------------------------------------------
    # 导入本地
    # ------------------------------------------------------------------
    def _find_or_create_account(self, name: str) -> Dict[str, Any]:
        accounts = self.portfolio_service.list_accounts(include_inactive=False)
        for a in accounts:
            if a["name"] == name:
                return a
        return self.portfolio_service.create_account(
            name=name,
            broker="ths",
            market="cn",
            base_currency="CNY",
        )

    def _import_trades(self, account_id: int, trades: List[Dict[str, Any]]) -> Dict[str, Any]:
        """按日期升序导入交易（幂等）。返回统计。"""
        sorted_trades = sorted(
            trades,
            key=lambda t: (str(t.get("trade_date") or ""), str(t.get("trade_time") or "")),
        )
        imported = 0
        skipped = 0
        errors: List[str] = []
        for t in sorted_trades:
            code = str(t.get("code") or "").strip()
            qty = t.get("quantity") or 0.0
            price = t.get("price") or 0.0
            side = t.get("side") or ""
            trade_date_str = str(t.get("trade_date") or "")
            if not code or _skip_symbol(code, str(t.get("name") or "")):
                skipped += 1
                continue
            if qty <= 0 or price <= 0 or side not in ("buy", "sell"):
                skipped += 1
                continue
            try:
                trade_date = date.fromisoformat(trade_date_str)
            except ValueError:
                skipped += 1
                continue
            try:
                self.portfolio_service.record_trade(
                    account_id=account_id,
                    symbol=code,
                    trade_date=trade_date,
                    side=side,
                    quantity=qty,
                    price=price,
                    fee=float(t.get("fee") or 0.0),
                    market="cn",
                    currency="CNY",
                    dedup_hash=_dedup_key(t),
                    note=t.get("name") or None,
                )
                imported += 1
            except Exception as exc:  # noqa: BLE001
                name = type(exc).__name__
                if "Conflict" in name or "duplicate" in str(exc).lower():
                    skipped += 1
                else:
                    errors.append(f"{code} {trade_date} {side}: {str(exc)[:80]}")
        return {"imported": imported, "skipped": skipped, "errors": errors[:10]}

    def _reconcile_positions(self, account_id: int, merged_positions: List[Dict[str, Any]]) -> Dict[str, Any]:
        """用账本汇总持仓校正本地持仓（以账本为准）：
        - 本地缺失 → 补初始建仓
        - 本地数量 < 账本 → 补差额买入
        - 本地数量 > 账本 → 补差额卖出
        """
        snapshot = self.portfolio_service.get_portfolio_snapshot(
            account_id=account_id,
            as_of=date.today(),
            cost_method="fifo",
            include_realtime=False,
        )
        local = {}
        for acc in snapshot.get("accounts", []):
            for p in acc.get("positions", []):
                local[str(p.get("symbol"))] = p
        added: List[str] = []
        adjusted: List[Dict[str, Any]] = []
        errors: List[str] = []
        for mp in merged_positions:
            code = str(mp.get("code") or "")
            qty = float(mp.get("quantity") or 0)
            if qty <= 0:
                continue
            cost = float(mp.get("cost") or 0) or 0.01
            name = str(mp.get("name") or "")
            lp = local.get(code)
            local_qty = float(lp.get("quantity") or 0) if lp else 0.0
            dedup_note = f"ths-reconcile|{code}|{qty}"
            try:
                if local_qty < qty:
                    diff = qty - local_qty
                    self.portfolio_service.record_trade(
                        account_id=account_id,
                        symbol=code,
                        trade_date=date.today(),
                        side="buy",
                        quantity=diff,
                        price=cost,
                        fee=0.0,
                        market="cn",
                        currency="CNY",
                        dedup_hash=hashlib.sha256(f"rec|{code}|{qty}|buy".encode("utf-8")).hexdigest(),
                        note=dedup_note,
                    )
                    if local_qty == 0:
                        added.append(code)
                    else:
                        adjusted.append({"code": code, "action": "add", "diff": diff})
                elif local_qty > qty:
                    diff = local_qty - qty
                    self.portfolio_service.record_trade(
                        account_id=account_id,
                        symbol=code,
                        trade_date=date.today(),
                        side="sell",
                        quantity=diff,
                        price=cost,
                        fee=0.0,
                        market="cn",
                        currency="CNY",
                        dedup_hash=hashlib.sha256(f"rec|{code}|{qty}|sell".encode("utf-8")).hexdigest(),
                        note=dedup_note,
                    )
                    adjusted.append({"code": code, "action": "reduce", "diff": diff})
            except Exception as exc:  # noqa: BLE001
                errors.append(f"{code}: {str(exc)[:80]}")
        return {"added": added, "adjusted": adjusted, "errors": errors[:10]}

    def _apply_ths_positions(self, account_id: int, merged_positions: List[Dict[str, Any]]) -> Dict[str, Any]:
        """用账本汇总持仓直接覆盖本地持仓（数量+成本以账本为准），成本精确。"""
        positions = []
        lots = []
        for mp in merged_positions:
            code = str(mp.get("code") or "")
            qty = float(mp.get("quantity") or 0)
            if qty <= 0:
                continue
            cost = float(mp.get("cost") or 0)
            last_price = float(mp.get("price") or 0) or cost
            positions.append({
                "symbol": code,
                "market": "cn",
                "currency": "CNY",
                "quantity": qty,
                "avg_cost": cost,
                "total_cost": round(qty * cost, 2),
                "last_price": last_price,
                "market_value_base": round(qty * last_price, 2),
                "unrealized_pnl_base": round(qty * (last_price - cost), 2),
                "valuation_currency": "CNY",
            })
            lots.append({
                "symbol": code,
                "market": "cn",
                "currency": "CNY",
                "open_date": date.today(),
                "remaining_quantity": qty,
                "unit_cost": cost,
                "source_trade_id": None,
            })
        self.repo.replace_positions_and_lots(
            account_id=account_id,
            cost_method="fifo",
            positions=positions,
            lots=lots,
            valuation_currency="CNY",
        )
        return {"applied": len(positions)}

    def _rebuild_ths_positions(self, account_id: int, merged_positions: List[Dict[str, Any]]) -> Dict[str, Any]:
        """重建持仓：清空该账户交易，为每只股票建一条账本成本的期初建仓（成本精确）。"""
        from sqlalchemy import delete as sa_delete
        from src.storage import PortfolioTrade

        with self.repo.db.get_session() as session:
            session.execute(sa_delete(PortfolioTrade).where(PortfolioTrade.account_id == account_id))
            session.commit()
        rebuilt: List[str] = []
        errors: List[str] = []
        for mp in merged_positions:
            code = str(mp.get("code") or "")
            qty = float(mp.get("quantity") or 0)
            cost = float(mp.get("cost") or 0)
            if qty <= 0 or cost <= 0:
                continue
            try:
                self.portfolio_service.record_trade(
                    account_id=account_id,
                    symbol=code,
                    trade_date=date.today(),
                    side="buy",
                    quantity=qty,
                    price=cost,
                    fee=0.0,
                    market="cn",
                    currency="CNY",
                    dedup_hash=hashlib.sha256(f"ths-base|{code}".encode("utf-8")).hexdigest(),
                    note=f"账本同步:{str(mp.get('name') or '')}",
                )
                rebuilt.append(code)
            except Exception as exc:  # noqa: BLE001
                errors.append(f"{code}: {str(exc)[:80]}")
        return {"rebuilt": rebuilt, "errors": errors[:10]}

    def _reconcile_cash(self, account_id: int, target_cash: float) -> Dict[str, Any]:
        """校正账户现金使其等于账本现金（幂等：先删旧校正记录再补差额）。"""
        # 1. 删除历史 ths现金 校正记录
        deleted = 0
        page = 1
        while True:
            page_result = self.portfolio_service.list_cash_ledger_events(
                account_id=account_id,
                page=page,
                page_size=100,
            )
            items = page_result.get("items", [])
            if not items:
                break
            for item in items:
                if (item.get("note") or "").startswith("ths现金"):
                    try:
                        self.portfolio_service.delete_cash_ledger_event(item["id"])
                        deleted += 1
                    except Exception:  # noqa: BLE001
                        pass
            if page * 100 >= page_result.get("total", 0):
                break
            page += 1
        # 2. 当前现金
        snapshot = self.portfolio_service.get_portfolio_snapshot(
            account_id=account_id,
            as_of=date.today(),
            cost_method="fifo",
            include_realtime=False,
        )
        current_cash = float(snapshot.get("accounts", [{}])[0].get("total_cash") or 0)
        diff = target_cash - current_cash
        if abs(diff) < 1.0:
            return {"adjusted": False, "current_cash": round(current_cash, 2), "deleted_old": deleted}
        direction = "in" if diff > 0 else "out"
        self.portfolio_service.record_cash_ledger(
            account_id=account_id,
            event_date=date.today(),
            direction=direction,
            amount=round(abs(diff), 2),
            note="ths现金校正",
        )
        return {
            "adjusted": True,
            "direction": direction,
            "amount": round(abs(diff), 2),
            "current_cash": round(current_cash, 2),
            "deleted_old": deleted,
        }

    def _import_asset_trend(self, account_id: int, points: List[Dict[str, Any]]) -> int:
        """把账本资产曲线写入本地日快照（仅 total_equity，用于资产曲线展示）。"""
        written = 0
        for p in points:
            d = str(p.get("date") or "")
            if not d:
                continue
            try:
                as_of = date.fromisoformat(d)
            except ValueError:
                continue
            equity = float(p.get("asset") or 0)
            if equity <= 0:
                continue
            try:
                self.repo.replace_positions_lots_and_snapshot(
                    account_id=account_id,
                    snapshot_date=as_of,
                    cost_method="fifo",
                    base_currency="CNY",
                    total_cash=0.0,
                    total_market_value=equity,
                    total_equity=equity,
                    unrealized_pnl=0.0,
                    realized_pnl=0.0,
                    fee_total=0.0,
                    tax_total=0.0,
                    fx_stale=False,
                    payload=json.dumps({"source": "ths_ledger", "equity_only": True}, ensure_ascii=False),
                    positions=[],
                    lots=[],
                    valuation_currency="CNY",
                )
                written += 1
            except Exception:  # noqa: BLE001
                continue
        return written

    # ------------------------------------------------------------------
    # 主入口
    # ------------------------------------------------------------------
    def sync(
        self,
        *,
        import_asset: bool = True,
        account_name: str = DEFAULT_ACCOUNT_NAME,
        start_date: Optional[str] = None,
        end_date: Optional[str] = None,
    ) -> Dict[str, Any]:
        if not self.client.is_logged_in():
            raise RuntimeError("同花顺账本未登录，请先扫码登录")
        data = self.fetch_merged(start_date=start_date, end_date=end_date)
        account = self._find_or_create_account(account_name)
        position_result = self._rebuild_ths_positions(account["id"], data["positions"])
        cash_result = self._reconcile_cash(account["id"], data["total_cash"])
        asset_points = []
        asset_written = 0
        if import_asset:
            try:
                asset_points = self.fetch_asset_trend_all()
                asset_written = self._import_asset_trend(account["id"], asset_points)
            except Exception:  # noqa: BLE001
                asset_written = 0
        return {
            "account": account,
            "accounts_scanned": len(data["accounts"]),
            "trades_fetched": data["trade_count"],
            "trades_imported": len(position_result["rebuilt"]),
            "positions_merged": data["position_count"],
            "positions_applied": len(position_result["rebuilt"]),
            "rebuilt": position_result["rebuilt"],
            "rebuild_errors": position_result["errors"],
            "cash_import": cash_result,
            "total_cash": data["total_cash"],
            "asset_points": len(asset_points),
            "asset_written": asset_written,
        }

    # ------------------------------------------------------------------
    # 对账：网页账本 vs 本地账户
    # ------------------------------------------------------------------
    def reconcile_web_vs_local(self, *, account_name: str = DEFAULT_ACCOUNT_NAME) -> Dict[str, Any]:
        """对账网页账本与本地账户，判断账目是否一致。

        对账口径：对比「网页汇总持仓」与「本地持仓」的 数量 + 成本（逐只）。
        市值差额单独展示（网页用账本实时价、本地用本地行情价，价格源不同会有小额差异，不判为账目错误）。
        现金差额作为参考展示。持仓数量或成本不一致时才提示导出核对。
        """
        if not self.client.is_logged_in():
            return {"available": False, "reason": "ths_not_logged_in", "aligned": None}

        # ---- 网页端：汇总持仓（数量/成本/市值） ----
        web = self.fetch_merged()
        web_cash = float(web.get("total_cash") or 0)
        web_positions = web.get("positions", [])
        web_by_code: Dict[str, Dict[str, Any]] = {}
        web_total_value = 0.0
        for p in web_positions:
            code = str(p.get("code") or "").strip()
            if not code:
                continue
            web_by_code[code] = {
                "name": str(p.get("name") or ""),
                "quantity": float(p.get("quantity") or 0),
                "cost": float(p.get("cost") or 0),
                "value": float(p.get("value") or 0),
            }
            web_total_value += float(p.get("value") or 0)

        # ---- 本地端：持仓（数量/成本/市值） ----
        account = self._find_or_create_account(account_name)
        snapshot = self.portfolio_service.get_portfolio_snapshot(
            account_id=account["id"], include_realtime=False
        )
        local_account = (snapshot.get("accounts") or [{}])[0]
        local_positions = local_account.get("positions") or []
        local_total_value = 0.0
        local_total_cash = float(local_account.get("total_cash") or 0)
        local_by_code: Dict[str, Dict[str, Any]] = {}
        for p in local_positions:
            code = str(p.get("symbol") or "").strip()
            if not code:
                continue
            local_by_code[code] = {
                "quantity": float(p.get("quantity") or 0),
                "cost": float(p.get("avg_cost") or 0),
                "value": float(p.get("market_value_base") or 0),
            }
            local_total_value += float(p.get("market_value_base") or 0)

        # ---- 逐只对比数量与成本 ----
        all_codes = sorted(set(web_by_code) | set(local_by_code))
        diff_positions: List[Dict[str, Any]] = []
        for code in all_codes:
            w = web_by_code.get(code)
            l = local_by_code.get(code)
            if w is None:
                diff_positions.append({"code": code, "issue": "网页有、本地无", "web": w, "local": l})
                continue
            if l is None:
                diff_positions.append({"code": code, "issue": "本地有、网页无", "web": w, "local": l})
                continue
            qty_ok = abs(w["quantity"] - l["quantity"]) <= 0.001
            cost_ok = abs(w["cost"] - l["cost"]) <= 0.005
            if not (qty_ok and cost_ok):
                diff_positions.append(
                    {
                        "code": code,
                        "issue": "数量或成本不一致",
                        "web": {"quantity": w["quantity"], "cost": round(w["cost"], 4)},
                        "local": {"quantity": l["quantity"], "cost": round(l["cost"], 4)},
                    }
                )

        positions_aligned = len(diff_positions) == 0
        value_diff = round(web_total_value - local_total_value, 2)
        cash_diff = round(web_cash - local_total_cash, 2)

        reasons: List[str] = []
        if not positions_aligned:
            reasons.append("持仓数量或成本与账本汇总持仓不一致")
            reasons.append("请导出 汇总持仓.xlsx 进行核对同步")
        elif abs(cash_diff) > 0.01:
            # 持仓一致但现金有差异（如银证转账），提示但不算账目错误
            reasons.append(
                f"持仓一致，但现金与账本存在差额 ¥{abs(cash_diff):,.2f}（可能是银证转账等，可导出文件核对）"
            )

        return {
            "available": True,
            "aligned": bool(positions_aligned),
            "position": {
                "web_count": len(web_positions),
                "local_count": len(local_by_code),
                "web_total_value": round(web_total_value, 2),
                "local_total_value": round(local_total_value, 2),
                "value_diff": value_diff,
                "diff_positions": diff_positions,
                "aligned": positions_aligned,
            },
            "cash": {
                "web": web_cash,
                "local": local_total_cash,
                "diff": cash_diff,
            },
            "reasons": reasons,
            "suggest_export": not positions_aligned,
        }

    # ------------------------------------------------------------------
    # 账本导出文件（汇总持仓.xlsx）导入
    # ------------------------------------------------------------------
    _EXPORT_CATEGORY_SPECIAL = ("银行转证券", "银行转存", "证券转银行", "银行转取")
    _EXPORT_CATEGORY_CASH_IN = ("银行转证券", "银行转存")
    _EXPORT_CATEGORY_CASH_OUT = ("证券转银行", "银行转取")
    _EXPORT_CATEGORY_TRADE = ("买入", "卖出")

    def parse_export_file(self, file_path: str) -> Dict[str, Any]:
        """解析账本导出的 汇总持仓.xlsx，返回持仓与交易记录统计（不写库）。"""
        import openpyxl

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        sheet_names = [str(s).strip() for s in wb.sheetnames]
        if "持仓数据" not in sheet_names:
            raise ValueError("导出文件缺少「持仓数据」工作表，请确认是账本导出的汇总持仓.xlsx")

        positions: List[Dict[str, Any]] = []
        funds_skipped: List[str] = []
        pos_ws = wb["持仓数据"]
        header = [str(c or "").strip() for c in next(pos_ws.iter_rows(values_only=True))]
        # 表头定位（兼容列序变化）
        def col_index(name: str) -> int:
            for i, h in enumerate(header):
                if name in h:
                    return i
            return -1

        idx_code = col_index("代码")
        idx_name = col_index("名称")
        idx_qty = col_index("持有数量")
        idx_cost = col_index("单位成本")

        # 精确列定位（避免「当日盈亏」误匹配「当日盈亏率」）
        def col_exact(name: str) -> int:
            for i, h in enumerate(header):
                if h == name:
                    return i
            return -1

        idx_price = col_exact("最新价")
        idx_value = col_exact("持有金额")
        idx_day = col_exact("当日盈亏")
        idx_day_pct = col_exact("当日盈亏率")
        idx_hold = col_exact("持有盈亏")

        def _row_float(i: int) -> Optional[float]:
            if i < 0:
                return None
            try:
                return float(row[i] or 0)
            except (TypeError, ValueError):
                return None

        summary: Dict[str, Any] = {}
        for row in pos_ws.iter_rows(min_row=2, values_only=True):
            code = str(row[idx_code] or "").strip() if idx_code >= 0 else ""
            name = str(row[idx_name] or "").strip() if idx_name >= 0 else ""
            if not code:
                continue
            if code == "汇总":
                summary = {
                    "total_market_value": _row_float(idx_value),
                    "day_pnl": _row_float(idx_day),
                    "day_pnl_pct": _row_float(idx_day_pct),
                    "hold_pnl": _row_float(idx_hold),
                }
                continue
            try:
                qty = float(row[idx_qty] or 0) if idx_qty >= 0 else 0.0
                cost = float(row[idx_cost] or 0) if idx_cost >= 0 else 0.0
            except (TypeError, ValueError):
                continue
            if _skip_symbol(code, name) or qty <= 0 or cost <= 0:
                continue
            if _is_fund_symbol(code, name):
                funds_skipped.append(f"{code} {name}".strip())
                continue
            positions.append(
                {
                    "code": code,
                    "name": name,
                    "quantity": qty,
                    "cost": cost,
                    "price": _row_float(idx_price),
                    "value": _row_float(idx_value),
                    "day_pnl": _row_float(idx_day),
                    "day_pnl_pct": _row_float(idx_day_pct),
                    "hold_pnl": _row_float(idx_hold),
                }
            )

        # 交易记录统计
        stats = {"trade_buy": 0, "trade_sell": 0, "cash_in": 0, "cash_out": 0, "other": 0, "first_date": None, "last_date": None}
        cash_in_total = 0.0
        cash_out_total = 0.0
        if "交易记录" in sheet_names:
            tr_ws = wb["交易记录"]
            tr_header = [str(c or "").strip() for c in next(tr_ws.iter_rows(values_only=True))]
            def tr_col(name: str) -> int:
                for i, h in enumerate(tr_header):
                    if name in h:
                        return i
                return -1
            ci_cat = tr_col("交易类别")
            ci_date = tr_col("成交日期")
            ci_money = tr_col("发生金额")
            for row in tr_ws.iter_rows(min_row=2, values_only=True):
                cat = str(row[ci_cat] or "").strip() if ci_cat >= 0 else ""
                if not cat:
                    continue
                d = str(row[ci_date] or "").strip() if ci_date >= 0 else ""
                if d and (stats["first_date"] is None or d < stats["first_date"]):
                    stats["first_date"] = d
                if d and (stats["last_date"] is None or d > stats["last_date"]):
                    stats["last_date"] = d
                if cat in self._EXPORT_CATEGORY_TRADE:
                    key = "trade_buy" if cat == "买入" else "trade_sell"
                    stats[key] += 1
                elif cat in self._EXPORT_CATEGORY_CASH_IN:
                    stats["cash_in"] += 1
                    try:
                        cash_in_total += abs(float(row[ci_money] or 0))
                    except (TypeError, ValueError):
                        pass
                elif cat in self._EXPORT_CATEGORY_CASH_OUT:
                    stats["cash_out"] += 1
                    try:
                        cash_out_total += abs(float(row[ci_money] or 0))
                    except (TypeError, ValueError):
                        pass
                else:
                    stats["other"] += 1

        total_mv = 0.0
        total_cost = 0.0
        for p in positions:
            total_mv += p["quantity"] * p["cost"]
            total_cost += p["quantity"] * p["cost"]
        return {
            "sheets": sheet_names,
            "position_count": len(positions),
            "positions": positions,
            "summary": summary,
            "funds_skipped": funds_skipped,
            "market_value": round(total_mv, 2),
            "trade_stats": stats,
            "cash_in_total": round(cash_in_total, 2),
            "cash_out_total": round(cash_out_total, 2),
        }

    def _parse_import_records(self, file_path: str) -> List[Dict[str, Any]]:
        """解析导出文件「交易记录」sheet 的全部流水（原类别原样），供持久化。"""
        import openpyxl
        from datetime import datetime as _dt

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        sheet_names = [str(s).strip() for s in wb.sheetnames]
        if "交易记录" not in sheet_names:
            return []
        ws = wb["交易记录"]
        header = [str(c or "").strip() for c in next(ws.iter_rows(values_only=True))]

        def idx(name: str) -> int:
            for i, h in enumerate(header):
                if name in h:
                    return i
            return -1

        i_date = idx("成交日期")
        i_time = idx("成交时间")
        i_code = idx("代码")
        i_name = idx("名称")
        i_cat = idx("交易类别")
        i_qty = idx("成交数量")
        i_price = idx("成交价格")
        i_amount = idx("发生金额")
        i_fee = idx("费用")
        i_note = idx("备注")
        records: List[Dict[str, Any]] = []
        seen: set = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            cat = str(row[i_cat] or "").strip() if i_cat >= 0 else ""
            if not cat:
                continue
            d = str(row[i_date] or "").strip() if i_date >= 0 else ""
            if not d:
                continue
            try:
                trade_date = _dt.strptime(d[:10], "%Y-%m-%d").date()
            except ValueError:
                continue
            code = str(row[i_code] or "").strip() if i_code >= 0 else ""
            name = str(row[i_name] or "").strip() if i_name >= 0 else ""
            try:
                qty = float(row[i_qty] or 0) if i_qty >= 0 else 0.0
                price = float(row[i_price] or 0) if i_price >= 0 else 0.0
                amount = float(row[i_amount] or 0) if i_amount >= 0 else 0.0
                fee = float(row[i_fee] or 0) if i_fee >= 0 else 0.0
            except (TypeError, ValueError):
                continue
            note = str(row[i_note] or "").strip() if i_note >= 0 else ""
            raw = "|".join([code, d, cat, str(qty), str(price), str(amount)])
            h = hashlib.sha256(raw.encode("utf-8")).hexdigest()
            if h in seen:
                continue
            seen.add(h)
            records.append({
                "record_type": cat,
                "symbol": code,
                "name": name,
                "trade_date": trade_date,
                "trade_time": str(row[i_time] or "").strip() if i_time >= 0 else "",
                "quantity": qty,
                "price": price,
                "amount": amount,
                "fee": fee,
                "note": note or None,
                "dedup_hash": h,
            })
        records.sort(key=lambda r: (r["trade_date"], r["trade_time"]))
        return records

    def import_export_file(
        self,
        file_path: str,
        *,
        account_name: str = DEFAULT_ACCOUNT_NAME,
        import_asset: bool = False,
    ) -> Dict[str, Any]:
        """把账本导出的 汇总持仓.xlsx 同步到本地账户。

        1) 持仓数据 sheet → 重建持仓（数量/成本与账本导出口径一致）
        2) 交易记录 sheet → **窗口覆盖**（只重写最近 WINDOW_TRADING_DAYS 个交易日，
           更早历史保留），含银证转账等全部类别，供报告展示
        3) 若账本已登录 → 现金校正；否则跳过现金校正
        """
        parsed = self.parse_export_file(file_path)
        if not parsed["positions"]:
            raise ValueError("导出文件未解析到有效持仓，请检查文件是否完整")
        account = self._find_or_create_account(account_name)
        position_result = self._rebuild_ths_positions(account["id"], parsed["positions"])
        # 完整交易流水持久化（含分红/银证转账等全部类别）——窗口覆盖，防止导出范围缩小误清历史
        import_records = self._parse_import_records(file_path)
        import_record_count = 0
        import_window = None
        import_kept_before = False
        import_warning = ""
        if import_records:
            # 本地已有最新交易日：窗口起点 = 本地最新交易日往前 WINDOW_TRADING_DAYS 个交易日；
            # 首次导入（本地无记录）：窗口起点 = 文件最早日期（全量写入）
            local_max = self.repo.max_import_trade_date(account["id"])
            file_max = max(r["trade_date"] for r in import_records)
            file_min = min(r["trade_date"] for r in import_records)
            if local_max is None:
                window_start = file_min
            else:
                window_start = _n_trading_days_before(local_max, WINDOW_TRADING_DAYS)
            # 兜底提示：文件最新日期早于本地最新，说明导出的不是最新
            if file_max < (local_max or date.today()):
                import_warning = (
                    f"导出文件最新记录仅到 {file_max.isoformat()}，本地已到 "
                    f"{local_max.isoformat()}，本次窗口仅覆盖最近{WINDOW_TRADING_DAYS}个交易日，"
                    f"请确认导出的是最新完整数据"
                )
            # 备份当前流水（保留最近一份，便于回滚）
            self.repo.backup_import_records(account["id"])
            win = self.repo.replace_import_records_window(
                account["id"], import_records, window_start
            )
            import_record_count = win["written"]
            import_window = win["window_start"]
            import_kept_before = win["kept_before_window"]
        cash_result: Dict[str, Any] = {}
        if self.client.is_logged_in():
            try:
                merged = self.fetch_merged()
                cash_result = self._reconcile_cash(account["id"], merged["total_cash"])
                total_cash = merged["total_cash"]
            except Exception as exc:  # noqa: BLE001
                cash_result = {"error": str(exc)[:120], "adjusted": False}
                total_cash = None
        else:
            cash_result = {"skipped": True, "note": "账本未登录，未做现金校正"}
            total_cash = None
        # 写入导出快照（今日盈亏/总市值/持有盈亏，供持仓页盘后读取）
        try:
            self._save_export_snapshot(parsed)
        except Exception:  # noqa: BLE001 - 快照是增强项，失败不阻断同步
            pass
        return {
            "account": account,
            "file": os.path.basename(file_path),
            "position_count": parsed["position_count"],
            "positions_applied": len(position_result["rebuilt"]),
            "funds_skipped": parsed.get("funds_skipped", []),
            "rebuilt": position_result["rebuilt"],
            "rebuild_errors": position_result["errors"],
            "trade_stats": parsed["trade_stats"],
            "cash_in_total": parsed["cash_in_total"],
            "cash_out_total": parsed["cash_out_total"],
            "market_value": parsed["market_value"],
            "cash_import": cash_result,
            "total_cash": total_cash,
            "import_record_count": import_record_count,
            "import_window_start": import_window,
            "import_kept_before": import_kept_before,
            "import_warning": import_warning,
        }

    def _save_export_snapshot(self, parsed: Dict[str, Any]) -> None:
        """把导出的账本汇总持仓快照写入本地 JSON，供持仓页盘后「今日盈亏」读取。

        仅当导出文件包含「汇总」行（总市值/当日盈亏/持有盈亏）时写入；
        快照以同步当天日期为 as_of，盘后且 as_of 为最近交易日时优先采用。
        """
        summary = parsed.get("summary") or {}
        if not summary.get("total_market_value") and not summary.get("day_pnl"):
            return
        payload = {
            "as_of": date.today().isoformat(),
            "day_pnl": summary.get("day_pnl"),
            "day_pnl_pct": summary.get("day_pnl_pct"),
            "total_market_value": summary.get("total_market_value"),
            "hold_pnl": summary.get("hold_pnl"),
            "positions": parsed.get("positions", []),
        }
        path = os.path.join(os.path.dirname(self.cookie_file), "ths_export_snapshot.json")
        with open(path, "w", encoding="utf-8") as fh:
            json.dump(payload, fh, ensure_ascii=False)

    # ------------------------------------------------------------------
    # 导出目录配置与自动检测（方案B：定时扫描下载目录）
    # ------------------------------------------------------------------
    def _export_config_path(self) -> str:
        return os.path.join(os.path.dirname(self.cookie_file), "ths_export_config.json")

    def get_export_config(self) -> Dict[str, Any]:
        cfg_path = self._export_config_path()
        config: Dict[str, Any] = {
            "directory": "",
            "auto_sync": False,
            "last_file": "",
            "last_synced_at": "",
            "last_sync_fingerprint": "",
        }
        if os.path.exists(cfg_path):
            try:
                with open(cfg_path, "r", encoding="utf-8") as fh:
                    config.update(json.load(fh))
            except Exception:  # noqa: BLE001
                pass
        return config

    def save_export_config(self, *, directory: Optional[str] = None, auto_sync: Optional[bool] = None) -> Dict[str, Any]:
        config = self.get_export_config()
        if directory is not None:
            config["directory"] = (directory or "").strip()
        if auto_sync is not None:
            config["auto_sync"] = bool(auto_sync)
        cfg_path = self._export_config_path()
        with open(cfg_path, "w", encoding="utf-8") as fh:
            json.dump(config, fh, ensure_ascii=False, indent=2)
        return config

    def find_latest_export(self, directory: Optional[str] = None) -> Optional[str]:
        """在指定目录（缺省用配置目录）找最新的 汇总持仓*.xlsx。"""
        config = self.get_export_config()
        directory = directory or config.get("directory") or ""
        if not directory or not os.path.isdir(directory):
            return None
        candidates: List[str] = []
        for fn in os.listdir(directory):
            low = fn.lower()
            if low.startswith("汇总持仓") and low.endswith(".xlsx") and not low.startswith("~$"):
                candidates.append(os.path.join(directory, fn))
        if not candidates:
            return None
        return max(candidates, key=lambda p: os.path.getmtime(p))

    @staticmethod
    def export_fingerprint(file_path: str) -> str:
        """生成导出文件指纹：文件名 + 修改时间 + 文件大小。

        用于自动检测时判断文件是否真的变化过（同名覆盖 / 加数字后缀都能正确识别）。
        """
        try:
            stat = os.stat(file_path)
            return "{}|{}|{}".format(
                os.path.basename(file_path),
                int(stat.st_mtime),
                int(stat.st_size),
            )
        except OSError:
            return ""

    def should_sync_export(self, directory: Optional[str] = None) -> Dict[str, Any]:
        """判断配置目录里最新的导出文件是否值得同步（指纹变化才需要同步）。

        返回 {"detected": bool, "latest": path|None, "changed": bool, "fingerprint": str}
        """
        latest = self.find_latest_export(directory=directory)
        if not latest:
            return {"detected": False, "latest": None, "changed": False, "fingerprint": ""}
        fingerprint = self.export_fingerprint(latest)
        config = self.get_export_config()
        last_fp = config.get("last_sync_fingerprint") or ""
        changed = bool(fingerprint) and fingerprint != last_fp
        return {
            "detected": True,
            "latest": latest,
            "changed": changed,
            "fingerprint": fingerprint,
        }

    def record_export_synced(self, file_path: str) -> None:
        """同步成功后记录本次文件的指纹与时间，供自动检测去重。"""
        config = self.get_export_config()
        cfg_path = self._export_config_path()
        now = __import__("datetime").datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        config.update(
            {
                "directory": config.get("directory", ""),
                "auto_sync": config.get("auto_sync", False),
                "last_file": os.path.basename(file_path),
                "last_synced_at": now,
                "last_sync_fingerprint": self.export_fingerprint(file_path),
            }
        )
        with open(cfg_path, "w", encoding="utf-8") as fh:
            json.dump(config, fh, ensure_ascii=False, indent=2)

    # ------------------------------------------------------------------
    # 独立访客分享页配置（展示股票白名单 + 固定口令）
    # ------------------------------------------------------------------
    def _share_config_path(self) -> str:
        return os.path.join(os.path.dirname(self.cookie_file), "share_config.json")

    def get_share_config(self) -> Dict[str, Any]:
        """分享页配置：{enabled, symbols, password_hash, password_salt, updated_at}。"""
        config: Dict[str, Any] = {
            "enabled": False,
            "symbols": [],
            "password_hash": "",
            "password_salt": "",
            "updated_at": "",
        }
        path = self._share_config_path()
        if os.path.exists(path):
            try:
                with open(path, "r", encoding="utf-8") as fh:
                    config.update(json.load(fh))
            except Exception:  # noqa: BLE001
                pass
        return config

    def save_share_config(self, config: Dict[str, Any]) -> Dict[str, Any]:
        import datetime
        config["updated_at"] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        path = self._share_config_path()
        with open(path, "w", encoding="utf-8") as fh:
            json.dump(config, fh, ensure_ascii=False, indent=2)
        return config

    @staticmethod
    def _hash_share_password(password: str, salt: str) -> str:
        return hashlib.sha256((salt + password).encode("utf-8")).hexdigest()

    def set_share_password(self, password: str) -> None:
        """设置固定口令（修改后旧口令立即失效）。传入空串则清除口令。"""
        cfg = self.get_share_config()
        password = (password or "").strip()
        if not password:
            cfg["password_hash"] = ""
            cfg["password_salt"] = ""
        else:
            salt = uuid.uuid4().hex
            cfg["password_hash"] = self._hash_share_password(password, salt)
            cfg["password_salt"] = salt
        self.save_share_config(cfg)

    def verify_share_password(self, password: str) -> bool:
        """校验固定口令；未设置口令或口令不匹配返回 False。"""
        cfg = self.get_share_config()
        h = cfg.get("password_hash") or ""
        salt = cfg.get("password_salt") or ""
        if not h or not salt:
            return False
        return h == self._hash_share_password(password or "", salt)

    def share_whitelist_symbols(self) -> List[str]:
        """返回分享页白名单股票代码（去空白、去重、保持顺序）。"""
        cfg = self.get_share_config()
        seen: List[str] = []
        for raw in cfg.get("symbols") or []:
            code = str(raw or "").strip()
            if code and code not in seen:
                seen.append(code)
        return seen

    # ------------------------------------------------------------------
    # 分享页访客会话（口令换取 12 小时有效会话，过期需重新输入口令）
    # ------------------------------------------------------------------
    def _share_session_path(self) -> str:
        return os.path.join(os.path.dirname(self.cookie_file), "share_session.json")

    def issue_share_session(self, hours: float = 12.0) -> str:
        """口令校验通过后发放一次性会话 token，存盘并返回。"""
        token = uuid.uuid4().hex + uuid.uuid4().hex
        payload = {
            "token": token,
            "issued_at": time.time(),
            "expires_at": time.time() + float(hours) * 3600,
        }
        path = self._share_session_path()
        try:
            with open(path, "w", encoding="utf-8") as fh:
                json.dump(payload, fh)
        except Exception:  # noqa: BLE001
            pass
        return token

    def verify_share_session(self, token: str) -> bool:
        """校验访客会话 token 是否有效（未过期）。"""
        if not token:
            return False
        try:
            with open(self._share_session_path(), "r", encoding="utf-8") as fh:
                data = json.load(fh)
            if data.get("token") != token:
                return False
            expires = float(data.get("expires_at") or 0)
            return time.time() < expires
        except Exception:  # noqa: BLE001
            return False


